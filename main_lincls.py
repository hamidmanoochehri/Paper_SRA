#!/usr/bin/env python
# original code obtained from https://github.com/facebookresearch/moco-v3
# Modified by Authors of SRA (paper link) 
# ------------------------------------------------------------------------------
# Copyright (c) Facebook, Inc. and its affiliates.
# All rights reserved.

# This source code is licensed under the license found in the
# LICENSE file in the root directory of this source tree.

import argparse
import builtins
import math
import os
import random
import shutil
import time
import warnings

import torch
import torch.nn as nn
import torch.nn.parallel
import torch.backends.cudnn as cudnn
import torch.distributed as dist
import torch.optim
import torch.multiprocessing as mp
import torch.utils.data
import torch.utils.data.distributed
import torchvision.transforms as transforms
import torchvision.datasets as datasets
import torchvision.models as torchvision_models
from  code_snapshotting import copy_files_with_timestamp
from datetime import datetime
import vits
import pandas as pd
from sklearn.metrics import classification_report, confusion_matrix, accuracy_score, balanced_accuracy_score, cohen_kappa_score, f1_score, recall_score, precision_score
import matplotlib.pyplot as plt
import warnings
from sklearn.exceptions import UndefinedMetricWarning
from sklearn.model_selection import KFold

from torch.utils.data.sampler import Sampler
from tensorboard.backend.event_processing.event_accumulator import EventAccumulator

import re
from openpyxl import load_workbook
from openpyxl import Workbook

from pytorch_balanced_sampler.sampler import SamplerFactory
import numpy as np
warnings.filterwarnings("ignore", category=UndefinedMetricWarning)

runs_root                   = './runs/'
code_snapshot_files         = ['classify_mocov3.sh', 'main_lincls.py', 'vits.py', 'code_snapshotting.py', 'moco/builder.py', 'moco/loader.py', 'moco/optimizer.py']

start_time                  = datetime.now()
run_timestamp               = start_time.strftime("%Y%m%d_%H%M%S")

torchvision_model_names = sorted(name for name in torchvision_models.__dict__
    if name.islower() and not name.startswith("__")
    and callable(torchvision_models.__dict__[name]))

model_names = ['vit_small', 'vit_base', 'vit_conv_small', 'vit_conv_base'] + torchvision_model_names

parser = argparse.ArgumentParser(description='PyTorch ImageNet Training')
parser.add_argument('--data', metavar='DIR',
                    help='path to dataset')
parser.add_argument('-a', '--arch', metavar='ARCH', default='resnet50',
                    choices=model_names,
                    help='model architecture: ' +
                        ' | '.join(model_names) +
                        ' (default: resnet50)')
parser.add_argument('-j', '--workers', default=32, type=int, metavar='N',
                    help='number of data loading workers (default: 32)')
parser.add_argument('--epochs', default=50, type=int, metavar='N',
                    help='number of total epochs to run')
parser.add_argument('--start-epoch', default=0, type=int, metavar='N',
                    help='manual epoch number (useful on restarts)')
parser.add_argument('-b', '--batch-size', default=1024, type=int,
                    metavar='N',
                    help='mini-batch size (default: 1024), this is the total '
                         'batch size of all GPUs on all nodes when '
                         'using Data Parallel or Distributed Data Parallel')
parser.add_argument('--use-cv', action='store_true',
                    help='Use cross-validation')
parser.add_argument('--num-folds', default=5, type=int,
                    help='Number of folds for cross-validation (default: 5)')
parser.add_argument('--lr', '--learning-rate', default=0.1, type=float,
                    metavar='LR', help='initial (base) learning rate', dest='lr')
parser.add_argument('--momentum', default=0.9, type=float, metavar='M',
                    help='momentum')
parser.add_argument('--wd', '--weight-decay', default=0., type=float,
                    metavar='W', help='weight decay (default: 0.)',
                    dest='weight_decay')
parser.add_argument('-p', '--print-freq', default=10, type=int,
                    metavar='N', help='print frequency (default: 10)')
parser.add_argument('--resume', default='', type=str, metavar='PATH',
                    help='path to latest checkpoint (default: none)')
parser.add_argument('-e', '--evaluate', dest='evaluate', action='store_true',
                    help='evaluate model on validation set')
parser.add_argument('--world-size', default=-1, type=int,
                    help='number of nodes for distributed training')
parser.add_argument('--rank', default=-1, type=int,
                    help='node rank for distributed training')
parser.add_argument('--dist-url', default='tcp://224.66.41.62:23456', type=str,
                    help='url used to set up distributed training')
parser.add_argument('--dist-backend', default='nccl', type=str,
                    help='distributed backend')
parser.add_argument('--seed', default=None, type=int,
                    help='seed for initializing training. ')
parser.add_argument('--gpu', default=None, type=int,
                    help='GPU id to use.')
parser.add_argument('--multiprocessing-distributed', action='store_true',
                    help='Use multi-processing distributed training to launch '
                         'N processes per node, which has N GPUs. This is the '
                         'fastest way to use PyTorch for either single node or '
                         'multi node data parallel training')

# additional configs:
parser.add_argument('--pretrained_folder', default='', type=str,
                    help='path to moco pretrained checkpoint')
parser.add_argument('--test-model', default='', type=str,
                    help='path to trained classifier checkpoint to be tested')
parser.add_argument('--test-data', default='', type=str, help='path to test dataset')
parser.add_argument('--test-mode', action='store_true', help='run in test mode')

# added
parser.add_argument('--iter_per_epoch', default=None, type=int,
                    help='manually set number of iters per epoch to have partial epochs (None or 0 for override).')
parser.add_argument('--balanced_sampler', action='store_true', 
                    help='whether to use balanced sampler or not.')

best_acc1 = 0

# ----------------------- find best pretr epoch from results folder path -----------------------
def find_best_epoch_from_tensorboard(tensorboard_log_dir):
    """
    This function parses the TensorBoard log files to find the epoch
    with the minimum validation loss.
    """
    # Initialize the event accumulator
    event_accumulator = EventAccumulator(tensorboard_log_dir)
    event_accumulator.Reload()

    # Extract the scalar events for the validation loss
    val_loss_events = event_accumulator.Scalars('val_loss')

    # Find the epoch with the minimum validation loss
    min_loss_epoch, min_loss = None, float('inf')
    for event in val_loss_events:
        if event.value < min_loss:
            min_loss = event.value
            min_loss_epoch = event.step

    return min_loss_epoch, min_loss
# --------------------------- for partial epochs manual_num_iters_per_epochs  ---------------------------
class BalancedSubsetSampler(Sampler):
    """Samples elements such that each class is equally represented up to num_samples_per_epoch"""
    def __init__(self, dataset, num_samples_per_epoch):
        self.dataset = dataset
        self.num_samples_per_epoch = num_samples_per_epoch
        self.class_counts = {}
        self.indices_per_class = {}
        
        for idx, (_, class_index) in enumerate(self.dataset):
            if class_index not in self.indices_per_class:
                self.indices_per_class[class_index] = []
            self.indices_per_class[class_index].append(idx)
        
        # We expect num_samples_per_epoch to be a multiple of the number of classes
        self.num_samples_per_class = self.num_samples_per_epoch // len(self.indices_per_class)

    def __iter__(self):
        indices = []
        for class_indices in self.indices_per_class.values():
            if len(class_indices) >= self.num_samples_per_class:
                indices += list(np.random.choice(class_indices, self.num_samples_per_class, replace=False))
            else:
                indices += class_indices + list(np.random.choice(class_indices, self.num_samples_per_class - len(class_indices), replace=True))
        np.random.shuffle(indices)
        return iter(indices)

    def __len__(self):
        return self.num_samples_per_epoch
# ---------------------------- test fn for additional testing and vis ------------------------------
def test(test_loader, model, args):
    batch_time = AverageMeter('Time', ':6.3f')
    top1 = AverageMeter('Acc@1', ':6.2f')
    top5 = AverageMeter('Acc@5', ':6.2f')
    progress = ProgressMeter(
        len(test_loader),
        [batch_time, top1, top5],
        prefix='Test: ')

    # switch to evaluate mode
    model.eval()
    all_preds = []
    all_targets = []

    with torch.no_grad():
        end = time.time()
        for i, (images, target) in enumerate(test_loader):
            if args.gpu is not None:
                images = images.cuda(args.gpu, non_blocking=True)
                target = target.cuda(args.gpu, non_blocking=True)
            if torch.cuda.is_available():
                target = target.cuda(args.gpu, non_blocking=True)

            # compute output
            output = model(images)
            _, preds = torch.max(output, 1)
            all_preds.extend(preds.cpu().numpy())
            all_targets.extend(target.cpu().numpy())

            # measure accuracy
            acc1, acc5 = accuracy(output, target, topk=(1, 5))
            top1.update(acc1[0], images.size(0))
            top5.update(acc5[0], images.size(0))

            # measure elapsed time
            batch_time.update(time.time() - end)
            end = time.time()

            if i % args.print_freq == 0:
                progress.display(i)

        print(' * Acc@1 {top1.avg:.3f} Acc@5 {top5.avg:.3f}'.format(top1=top1, top5=top5))

    # Classification metrics and report
    y_true = all_targets
    y_pred = all_preds
    conf_matrix = confusion_matrix(y_true, y_pred)
    balanced_acc = balanced_accuracy_score(y_true, y_pred)
    class_accs = accuracy_score(y_true, y_pred)

    # Assuming get_save_folder_path function exists and provides the correct output directory
    save_path = get_save_folder_path(args)
    print(f'Saving classification test report to {save_path}')
    compute_and_save_classification_report(y_true, y_pred, save_path, "test", conf_matrix, args)

    return top1.avg, balanced_acc, class_accs, all_targets, all_preds, conf_matrix

def parse_classify_script(script_path):
    # Initialize a dictionary to hold the parameters
    params = {}

    with open(script_path, 'r') as file:
        lines = file.readlines()
    
    # Concatenate lines to process them as a single string
    script_content = " ".join(line.strip() for line in lines if not line.strip().startswith('#'))
    #print(script_content)
    # Search for the CUDA_VISIBLE_DEVICES line and subsequent arguments
    cmd_start = script_content.find('CUDA_VISIBLE_DEVICES')
    if cmd_start == -1:
        print("CUDA_VISIBLE_DEVICES not found in the script.")
        return params

    # Extract arguments using regular expressions
    params['lr'] = re.search(r'--lr (\d+\.\d+)', script_content).group(1) if re.search(r'--lr (\d+\.\d+)', script_content) else None
    params['arch'] = re.search(r'--arch (\w+)', script_content).group(1) if re.search(r'--arch (\w+)', script_content) else None
    params['batch_size'] = re.search(r'--batch-size (\d+)', script_content).group(1) if re.search(r'--batch-size (\d+)', script_content) else None
    data_match = re.search(r'--data ([^\s]+)', script_content)
    if data_match:
        data_path = data_match.group(1)
        # Check the entire path for keywords
        if 'ukidney' in data_path or 'utah_kidney' in data_path:
            params['tiles'] = 'ukidney'
        elif 'tcga' in data_path:
            params['tiles'] = 'tcga'
        else:
            params['tiles'] = data_path.split('/')[-1]
    pretr_epoch_match = re.search(r'checkpoint_(\d+).pth.tar', script_content)
    params['pretr_epoch'] = pretr_epoch_match.group(1).lstrip('0') if pretr_epoch_match else None

    # Mapping architecture to shortened form
    if params.get('arch') == 'resnet50':
        params['arch'] = 'rn50'
    elif params.get('arch') == 'resnet18':
        params['arch'] = 'rn18'
    
    print(params)
    return params

class CustomImageFolder(datasets.ImageFolder):
    def __init__(self, root, transform=None):
        super(CustomImageFolder, self).__init__(root, transform=transform)
        self.class_idxs = [[] for _ in range(len(self.classes))]

        # Organize indices by class
        for idx, (_, class_idx) in enumerate(self.samples):
            self.class_idxs[class_idx].append(idx)
    
    def __getitem__(self, index):
        path, target = self.samples[index]
        sample = self.loader(path)
        if self.transform is not None:
            sample = self.transform(sample)
        if self.target_transform is not None:
            target = self.target_transform(target)

        return sample, target, path  # return path for each sample
    
# ---------------------------- plot accuracies fn ------------------------------
def plot_accuracies(train_accuracies, val_accuracies, output_directory, args):
    plt.figure(figsize=(14, 6))

    # Subplot 1: Train vs Val Accuracy
    plt.subplot(1, 2, 1)
    plt.plot(range(1, args.epochs + 1), train_accuracies, label='Train Accuracy')
    plt.plot(range(1, args.epochs + 1), val_accuracies, label='Validation Accuracy')
    plt.title('Training and Validation Accuracy over Epochs')
    plt.xlabel('Epochs')
    plt.ylabel('Accuracy')
    plt.legend()

    # Subplot 2: Per-class Train vs Val Accuracies
    # You need to modify this part to include your per-class accuracies
    plt.subplot(1, 2, 2)
    # plt.plot(...) # Add plots for per-class accuracies here
    plt.title('Per-Class Training and Validation Accuracy over Epochs')
    plt.xlabel('Epochs')
    plt.ylabel('Accuracy')
    plt.legend()

    plt.tight_layout()
    plt.savefig(os.path.join(output_directory, 'accuracy_plots.png'))
    plt.show()
# ------------------------------------------------------------------------------

# ----------------------- classification metrics and report --------------------
def compute_and_save_classification_report(y_true, y_pred, output_directory, epoch, conf_matrix, args):
    # Compute metrics
    report_dict = classification_report(y_true, y_pred, output_dict=True, zero_division=0)
    balanced_acc = balanced_accuracy_score(y_true, y_pred)

    # Per-Class Accuracy
    per_class_accuracy = conf_matrix.diagonal() / conf_matrix.sum(axis=1)

    # Process per-class metrics and add additional details
    class_report = []
    for i, class_id in enumerate(report_dict.keys()):
        if class_id in ['0', '1', '2', '3']:
            row = {
                'epoch': epoch,
                'class': class_id,
                'accuracy': per_class_accuracy[i],
                'true_numbers': report_dict[class_id]['support'],
                'confusion_matrix_0': conf_matrix[int(class_id)][0],  # Adjust indices based on your class labels
                'confusion_matrix_1': conf_matrix[int(class_id)][1],
                'confusion_matrix_2': conf_matrix[int(class_id)][2],
                #'confusion_matrix_3': conf_matrix[int(class_id)][3],
                'precision': report_dict[class_id]['precision'],
                'recall': report_dict[class_id]['recall'],
                'f1-score': report_dict[class_id]['f1-score']
            }
            class_report.append(row)

    # Compute metrics
    report_dict = classification_report(y_true, y_pred, output_dict=True, zero_division=0)
    accuracy = accuracy_score(y_true, y_pred)
    balanced_acc = balanced_accuracy_score(y_true, y_pred)
    kappa = cohen_kappa_score(y_true, y_pred)
    macro_f1 = f1_score(y_true, y_pred, average='macro')
    weighted_f1 = f1_score(y_true, y_pred, average='weighted')
    macro_precision = precision_score(y_true, y_pred, average='macro', zero_division=0)
    weighted_precision = precision_score(y_true, y_pred, average='weighted', zero_division=0)
    macro_recall = recall_score(y_true, y_pred, average='macro', zero_division=0)
    weighted_recall = recall_score(y_true, y_pred, average='weighted', zero_division=0)

    # Aggregate metrics
    summary_metrics = {
        'epoch': epoch,
        'accuracy': accuracy,
        'balanced_accuracy': balanced_acc,
        'cohen_kappa': kappa,
        'macro_f1_score': macro_f1,
        'weighted_f1_score': weighted_f1,
        'macro_precision': macro_precision,
        'weighted_precision': weighted_precision,
        'macro_recall': macro_recall,
        'weighted_recall': weighted_recall
    }

    class_report_df = pd.DataFrame(class_report)

    checkpoint_base_name = os.path.basename(args.test_model)
    checkpoint_number, _ = os.path.splitext(os.path.splitext(checkpoint_base_name)[0])
    if args.test_mode: # indicating test
        summary_csv_path = os.path.join(output_directory, f'summary_metrics_test_{checkpoint_number}.csv')
        conf_matrix_path = os.path.join(output_directory, f'conf_matrix_test_{checkpoint_number}.csv')
        report_csv_path = os.path.join(output_directory, f'classification_report_test_{checkpoint_number}.csv')
    else:
        summary_csv_path = os.path.join(output_directory, 'summary_metrics.csv')
        conf_matrix_path = os.path.join(output_directory, 'conf_matrix.csv')
        report_csv_path = os.path.join(output_directory, 'classification_report.csv')
    class_report_df.to_csv(report_csv_path, mode='a', header=not os.path.exists(report_csv_path), index=False)
    
    conf_matrix_df   = pd.DataFrame(conf_matrix)
    summary_df       = pd.DataFrame([summary_metrics])
    #print(f'*** conf matrix elements: conf_matrix[0][0]: {conf_matrix[0][0]}, {conf_matrix[0][1]},  {conf_matrix[0][2]},  {conf_matrix[0][3]}')
    summary_df.to_csv(summary_csv_path, mode='a', header=not os.path.exists(summary_csv_path), index=False)
    conf_matrix_df.to_csv(conf_matrix_path)

    #print(f"Appended classification report to {report_csv_path}")
    #print(f"Appended summary metrics to {summary_csv_path}")


    #print(f"Saved classification report to {report_csv_path}")
    #print(f"Saved confusion matrix to {conf_matrix_csv_path}")
    #print(f"Saved summary metrics to {summary_csv_path}")
    
    # New: Return values for Excel export
    return {
        'conf_matrix': conf_matrix,
        'class_report': class_report,
        'summary_metrics': summary_metrics
    }
# ------------------------------------------------------------------------------

def extract_info_from_path(test_model_path):
    parts = test_model_path.strip(os.path.sep).split(os.path.sep)
    pretr_exp_number = parts[-5]
    pretr_checkpoint = parts[-3].split('_')[1]
    cls_exp_number = parts[-2]
    cls_checkpoint = parts[-1].split('_')[-1].split('.')[0] 

    return pretr_exp_number, pretr_checkpoint, cls_exp_number, cls_checkpoint

# New function to write results to an Excel file
def write_to_excel(params, conf_matrix, class_report, summary_metrics, args):
    excel_file_path = 'all_test_results.xlsx'
    pretr_exp_number, pretr_checkpoint, cls_exp_number, cls_checkpoint = extract_info_from_path(args.test_model)
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"Excel file does not exist: {excel_file_path}")

    # Load the workbook
    workbook = load_workbook(excel_file_path)
    
    # Check if 'Sheet1' exists
    if 'Sheet1' not in workbook.sheetnames:
        raise ValueError(f"'Sheet1' does not exist in {excel_file_path}")

    # Initialize ExcelWriter with the loaded workbook
    writer = pd.ExcelWriter(excel_file_path, engine='openpyxl')
    writer.book = workbook
    writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)

    # Determine the starting row for new data
    startrow = writer.sheets['Sheet1'].max_row if writer.sheets['Sheet1'].max_row > 1 else 1

    # Prepare pretraining and classification info for writing
    info_df = pd.DataFrame([{
        'pretr_exp_number': pretr_exp_number,
        'pretr_checkpoint': pretr_checkpoint,
        'cls_exp_number': cls_exp_number,
        'cls_checkpoint': cls_checkpoint
    }])
    
    details_df = pd.DataFrame([params])

    params.update({
        'pretr_exp_number': pretr_exp_number,
        'pretr_checkpoint': pretr_checkpoint,
        'cls_exp_number': cls_exp_number,
        'cls_checkpoint': cls_checkpoint
    })

    # Prepare dataframes from provided data
    conf_matrix_df = pd.DataFrame(conf_matrix)
    class_report_df = pd.DataFrame(class_report)
    class_report_df.drop(['epoch','confusion_matrix_0', 'confusion_matrix_1', 'confusion_matrix_2'], axis=1, inplace=True)
    # Ensure summary_metrics is not including 'epoch' column
    summary_metrics_df = pd.DataFrame([summary_metrics])
    summary_metrics_df.drop('epoch', axis=1, inplace=True)
    combined_df = pd.concat([info_df, details_df], axis=1)
    combined_df = combined_df[['pretr_exp_number', 'pretr_checkpoint', 'cls_exp_number', 'cls_checkpoint', 'tiles', 'batch_size', 'arch', 'lr', 'pretr_epoch']]
    conf_matrix_df['class'] = class_report_df['class']
    conf_matrix_df['true_numbers'] = class_report_df['true_numbers']
    conf_matrix_columns = ['class'] + list(range(conf_matrix.shape[0])) + ['true_numbers']
    conf_matrix_df = conf_matrix_df[conf_matrix_columns]
    class_report_df.drop(['class','true_numbers'], axis=1, inplace=True)
    combined_df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, index=False, header=False)
    conf_matrix_df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, startcol=8, index=False, header=False)
    class_report_df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, startcol=14, index=False, header=False)
    summary_metrics_df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, startcol=18, index=False, header=False)

    writer.save()

def get_save_folder_path(args):
    if False:
        # determine save dir
        checkpoint_filename    = os.path.basename(args.pretrained_folder)  # Get the filename from the path
        #print(f'*** checkpoint_filename: {checkpoint_filename.split(".")[0]}')
        checkpoint_number      = checkpoint_filename.split('.')[0]  # Remove the file extension
        checkpoint_run_name    = args.pretrained_folder.split("/")[-3]

        if args.test_mode:
            save_root = os.path.dirname(args.test_model)
        else:
            save_root = os.path.join(runs_root, checkpoint_run_name, 'classification_results', checkpoint_number, run_timestamp)

    checkpoint_number , _ = find_best_epoch_from_tensorboard(args.pretrained_folder)
    save_root = os.path.join(args.pretrained_folder, 'classification_results', str(checkpoint_number), run_timestamp)
    return save_root

# Define normalize globally
normalize = transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])

def main():
    args = parser.parse_args()

    # getting snapshot of the codes
    if not args.test_mode:
        copy_files_with_timestamp(code_snapshot_files, save_root=get_save_folder_path(args), run_timestamp=run_timestamp) # take source snapshot

    if args.seed is not None:
        random.seed(args.seed)
        torch.manual_seed(args.seed)
        cudnn.deterministic = True
        warnings.warn('You have chosen to seed training. '
                      'This will turn on the CUDNN deterministic setting, '
                      'which can slow down your training considerably! '
                      'You may see unexpected behavior when restarting '
                      'from checkpoints.')

    if args.gpu is not None:
        warnings.warn('You have chosen a specific GPU. This will completely '
                      'disable data parallelism.')

    if args.dist_url == "env://" and args.world_size == -1:
        args.world_size = int(os.environ["WORLD_SIZE"])

    args.distributed = args.world_size > 1 or args.multiprocessing_distributed
    print(f'*** args.distributed: {args.distributed}')

    ngpus_per_node = torch.cuda.device_count()

    if args.test_mode:
        # Running in test mode, only one GPU will be used in this case
        main_worker(args.gpu if args.gpu is not None else 0, ngpus_per_node, args)

    else:
        if args.multiprocessing_distributed:
            # Since we have ngpus_per_node processes per node, the total world_size
            # needs to be adjusted accordingly
            args.world_size = ngpus_per_node * args.world_size
            # Use torch.multiprocessing.spawn to launch distributed processes: the
            # main_worker process function
            mp.spawn(main_worker, nprocs=ngpus_per_node, args=(ngpus_per_node, args))
        else:
            # Simply call main_worker function
            main_worker(args.gpu, ngpus_per_node, args)


def main_worker(gpu, ngpus_per_node, args):
    global best_acc1
    args.gpu = gpu
    
    normalize = transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])

    # suppress printing if not master
    if args.multiprocessing_distributed and args.gpu != 0:
        def print_pass(*args):
            pass
        builtins.print = print_pass

    if args.gpu is not None:
        print("Use GPU: {} for training".format(args.gpu))

    if args.distributed:
        if args.dist_url == "env://" and args.rank == -1:
            args.rank = int(os.environ["RANK"])
        if args.multiprocessing_distributed:
            # For multiprocessing distributed training, rank needs to be the
            # global rank among all the processes
            args.rank = args.rank * ngpus_per_node + gpu
        dist.init_process_group(backend=args.dist_backend, init_method=args.dist_url,
                                world_size=args.world_size, rank=args.rank)
        torch.distributed.barrier()

    print(f'Run Number (run_timestamp): {run_timestamp}')

    # create model
    print("=> creating model '{}'".format(args.arch))
    if args.arch.startswith('vit'):
        model = vits.__dict__[args.arch]()
        linear_keyword = 'head'
    else:
        model = torchvision_models.__dict__[args.arch]()
        linear_keyword = 'fc'

    # freeze all layers but the last fc
    for name, param in model.named_parameters():
        if name not in ['%s.weight' % linear_keyword, '%s.bias' % linear_keyword]:
            param.requires_grad = False
    # init the fc layer
    getattr(model, linear_keyword).weight.data.normal_(mean=0.0, std=0.01)
    getattr(model, linear_keyword).bias.data.zero_()

    # ---------------------- for additional testing and vis ----------------------
    # Load and preprocess test data
    if args.test_mode:
        testdir = args.test_data
        test_loader = torch.utils.data.DataLoader(
            datasets.ImageFolder(testdir, transforms.Compose([
                transforms.Resize(256),
                transforms.CenterCrop(224),
                transforms.ToTensor(),
                normalize,
            ])),
            batch_size=args.batch_size, shuffle=False,
            num_workers=args.workers, pin_memory=True
        )
        
        if args.test_model and os.path.isfile(args.test_model):
            print(f"=> loading test model '{args.test_model}'")
            checkpoint = torch.load(args.test_model, map_location="cpu")
            
            # Check if model was saved with DataParallel, which prefixes state dict keys with 'module.'
            if list(checkpoint['state_dict'].keys())[0].startswith('module.'):
                # Create a new state dict without the 'module.' prefix
                state_dict = {key.replace('module.', ''): value for key, value in checkpoint['state_dict'].items()}
            else:
                state_dict = checkpoint['state_dict']

            # Load adjusted state dict
            model.load_state_dict(state_dict)
            print("=> loaded test model")

        if torch.cuda.is_available():
            print("=> using GPU for testing")
            model = model.cuda(args.gpu)
        else:
            print("=> using CPU for testing")

        save_path = get_save_folder_path(args)

        test_acc, balanced_acc, class_accs, y_true, y_pred, conf_matrix = test(test_loader, model, args)
        print(f"Test Accuracy: {test_acc:.2f}, Balanced Accuracy: {balanced_acc:.4f}")
        results = compute_and_save_classification_report(y_true, y_pred, save_path, "test", conf_matrix, args)
        
        # Write to Excel
        write_to_excel(results['conf_matrix'], results['class_report'], results['summary_metrics'], args)
    # ----------------------------------------------------------------------------
    if not args.test_mode:
        # Data loading code
        print(f'data folder: {args.data}')
        traindir = os.path.join(args.data, 'train')
        valdir = os.path.join(args.data, 'val')
        # load from pre-trained, before DistributedDataParallel constructor
        if args.pretrained_folder:
            best_epoch, min_loss = find_best_epoch_from_tensorboard(args.pretrained_folder)
            print(f"Best epoch: {best_epoch} with minimum validation loss: {min_loss}")
            pretrained_file = os.path.join(args.pretrained_folder, 'saved_checkpoints', f'checkpoint_{best_epoch:04}.pth.tar')
            if os.path.isfile(pretrained_file):
                print("=> loading checkpoint '{}'".format(pretrained_file))
                checkpoint = torch.load(pretrained_file, map_location="cpu")
                #print(f' *** checkpoint keys: {checkpoint.keys()}')
                # rename moco pre-trained keys
                state_dict = checkpoint['state_dict'] # pretrained by moco: 'state_dict', pretrained by dino: 'student'
                # show student keys:
                #print("Keys in 'student' state_dict:")
                #for key in state_dict.keys():
                #    print(key)
                for k in list(state_dict.keys()):
                    # retain only base_encoder up to before the embedding layer
                    if k.startswith('module.base_encoder') and not k.startswith('module.base_encoder.%s' % linear_keyword):
                        # remove prefix
                        state_dict[k[len("module.base_encoder."):]] = state_dict[k]
                    # delete renamed or unused k
                    del state_dict[k]

                args.start_epoch = 0
                msg = model.load_state_dict(state_dict, strict=False)
                assert set(msg.missing_keys) == {"%s.weight" % linear_keyword, "%s.bias" % linear_keyword}

                print("=> loaded pre-trained model '{}'".format(pretrained_file))
            else:
                print("=> no checkpoint found at '{}'".format(pretrained_file))

        # infer learning rate before changing batch size
        init_lr = args.lr * args.batch_size / 256

        if not torch.cuda.is_available():
            print('using CPU, this will be slow')
        elif args.distributed:
            # For multiprocessing distributed, DistributedDataParallel constructor
            # should always set the single device scope, otherwise,
            # DistributedDataParallel will use all available devices.
            if args.gpu is not None:
                torch.cuda.set_device(args.gpu)
                model.cuda(args.gpu)
                # When using a single GPU per process and per
                # DistributedDataParallel, we need to divide the batch size
                # ourselves based on the total number of GPUs we have
                args.batch_size = int(args.batch_size / args.world_size)
                args.workers = int((args.workers + ngpus_per_node - 1) / ngpus_per_node)
                model = torch.nn.parallel.DistributedDataParallel(model, device_ids=[args.gpu])
            else:
                model.cuda()
                # DistributedDataParallel will divide and allocate batch_size to all
                # available GPUs if device_ids are not set
                model = torch.nn.parallel.DistributedDataParallel(model)
        elif args.gpu is not None:
            torch.cuda.set_device(args.gpu)
            model = model.cuda(args.gpu)
        else:
            # DataParallel will divide and allocate batch_size to all available GPUs
            if args.arch.startswith('alexnet') or args.arch.startswith('vgg'):
                model.features = torch.nn.DataParallel(model.features)
                model.cuda()
            else:
                model = torch.nn.DataParallel(model).cuda()

        # define loss function (criterion) and optimizer
        criterion = nn.CrossEntropyLoss().cuda(args.gpu)

        # optimize only the linear classifier
        parameters = list(filter(lambda p: p.requires_grad, model.parameters()))
        assert len(parameters) == 2  # weight, bias

        optimizer = torch.optim.SGD(parameters, init_lr,
                                    momentum=args.momentum,
                                    weight_decay=args.weight_decay)
    
        # optionally resume from a checkpoint
        if args.resume:
            if os.path.isfile(args.resume):
                print("=> loading checkpoint '{}'".format(args.resume))
                if args.gpu is None:
                    checkpoint = torch.load(args.resume)
                else:
                    # Map model to be loaded to specified single gpu.
                    loc = 'cuda:{}'.format(args.gpu)
                    checkpoint = torch.load(args.resume, map_location=loc)
                args.start_epoch = checkpoint['epoch']
                best_acc1 = checkpoint['best_acc1']
                if args.gpu is not None:
                    # best_acc1 may be from a checkpoint from a different GPU
                    best_acc1 = best_acc1.to(args.gpu)
                model.load_state_dict(checkpoint['state_dict'])
                optimizer.load_state_dict(checkpoint['optimizer'])
                print("=> loaded checkpoint '{}' (epoch {})"
                    .format(args.resume, checkpoint['epoch']))
            else:
                print("=> no checkpoint found at '{}'".format(args.resume))

        cudnn.benchmark = True

        print(f'data folder: {args.data}')

        normalize = transforms.Normalize(mean=[0.485, 0.456, 0.406],
                                        std=[0.229, 0.224, 0.225])


        # Setup the training dataset and DataLoader
        train_dataset = CustomImageFolder(
            os.path.join(args.data, 'train'),
            transforms.Compose([
                transforms.RandomResizedCrop(224),
                transforms.RandomHorizontalFlip(),
                transforms.ToTensor(),
                normalize,
            ])
        )

        if args.iter_per_epoch and args.balanced_sampler:
            iter_per_epoch = args.iter_per_epoch
        elif args.balanced_sampler and not args.iter_per_epoch:
            iter_per_epoch = len(train_dataset) // args.batch_size + (len(train_dataset) % args.batch_size != 0)
        

        #https://github.com/khornlund/pytorch-balanced-sampler
        batch_sampler_labeled = SamplerFactory().get(
            class_idxs = train_dataset.class_idxs,#index start with 0: [[0,1,2,3],[4,5,6],[7],[8],[9]], 5 classes
            batch_size = args.batch_size,
            n_batches  = iter_per_epoch,#how many batches in one epoch
            alpha      = 1.0,#totally balanced
            kind       = 'fixed'#fixed number of each label in one batch
            )

        # Use the custom sampler in the DataLoader
        train_loader = torch.utils.data.DataLoader(
            train_dataset,
            batch_sampler=batch_sampler_labeled,
            num_workers=args.workers,
            pin_memory=True
        )

        
        if args.distributed:
            train_sampler = torch.utils.data.distributed.DistributedSampler(train_dataset)
        else:
            train_sampler = None

        # Using standard ImageFolder for standard non-balanced sampler
        train_loader_ = torch.utils.data.DataLoader(
            train_dataset, batch_size=args.batch_size, shuffle=(train_sampler is None),
            num_workers=args.workers, pin_memory=True, sampler=train_sampler)

        val_loader = torch.utils.data.DataLoader(
            datasets.ImageFolder(valdir, transforms.Compose([
                transforms.Resize(256),
                transforms.CenterCrop(224),
                transforms.ToTensor(),
                normalize,
            ])),
            batch_size=256, shuffle=False,
            num_workers=args.workers, pin_memory=True)

        if args.evaluate:
            validate(val_loader, model, criterion, args)
            return

        train_balanced_accuracies = []
        val_balanced_accuracies   = []
        train_class_accuracies    = []  # This should be a list of lists, one for each class
        val_class_accuracies      = []  # Similarly, a list of lists for validation

        for epoch in range(args.start_epoch, args.epochs):
            if args.distributed:
                train_sampler.set_epoch(epoch)
            adjust_learning_rate(optimizer, init_lr, epoch, args)

            # train for one epoch
            train(train_loader, model, criterion, optimizer, epoch, args)

            # evaluate on validation set
            #acc1 = validate(val_loader, model, criterion, args)
            # Evaluate on the validation set and unpack the returned tuple
            val_metrics = validate(val_loader, model, criterion, args, epoch=epoch)
            acc1 = val_metrics[0]  # Assuming the first element is the regular accuracy

            # remember best acc@1 and save checkpoint
            is_best = acc1 > best_acc1
            best_acc1 = max(acc1, best_acc1)

            if not args.multiprocessing_distributed or (args.multiprocessing_distributed
                    and args.rank == 0): # only the first GPU saves checkpoint
                
                checkpoint_save_folder = get_save_folder_path(args)
                save_checkpoint({
                    'epoch': epoch + 1,
                    'arch': args.arch,
                    'state_dict': model.state_dict(),
                    'best_acc1': best_acc1,
                    'optimizer' : optimizer.state_dict(),
                }, is_best, checkpoint_save_folder, f'checkpoint_{epoch}.pth.tar')
                if epoch == args.start_epoch:
                    sanity_check(model.state_dict(), pretrained_file, linear_keyword)


def train(train_loader, model, criterion, optimizer, epoch, args):
    #print(f'pretrained model: {args.pretrained}')
    batch_time = AverageMeter('Time', ':6.3f')
    data_time = AverageMeter('Data', ':6.3f')
    losses = AverageMeter('Loss', ':.4e')
    top1 = AverageMeter('Acc@1', ':6.2f')
    top5 = AverageMeter('Acc@5', ':6.2f')
    progress = ProgressMeter(
        len(train_loader),
        [batch_time, data_time, losses, top1, top5],
        prefix="Epoch: [{}]".format(epoch))

    model.eval()  # Ensure the model is in evaluation mode

    end = time.time()
    for i, (images, targets, paths) in enumerate(train_loader):  # Ensure loader returns paths
        # Measure data loading time
        data_time.update(time.time() - end)

        if args.gpu is not None:
            images = images.cuda(args.gpu, non_blocking=True)
        if torch.cuda.is_available():
            targets = targets.cuda(args.gpu, non_blocking=True)

        # Compute output
        output = model(images)
        loss = criterion(output, targets)

        _, predicted = torch.max(output, 1)

        # Measure accuracy and record loss
        acc1, acc5 = accuracy(output, targets, topk=(1, 5))
        losses.update(loss.item(), images.size(0))
        top1.update(acc1[0], images.size(0))
        top5.update(acc5[0], images.size(0))

        # Compute gradient and do SGD step
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()

        # Measure elapsed time
        batch_time.update(time.time() - end)
        end = time.time()

        # Optionally print the batch details to a text file
        #if i == 0:  # Example: write details for the first batch of each epoch
        #    batch_data = list(zip(paths, targets.cpu().numpy()))
        #    write_batch_details(batch_data, f'batch_details_epoch_{epoch}.txt', i)

        if i % args.print_freq == 0:
            progress.display(i)


def write_batch_details(batch_data, filepath, batch_idx):
    from collections import Counter

    # Count the number of data points per class
    class_counts = Counter([label for _, label in batch_data])

    with open(filepath, 'a') as f:
        f.write(f'Batch {batch_idx} - Class Counts: {dict(class_counts)}\n')
        for image_path, label in batch_data:
            f.write(f'{image_path} {label}\n')
        f.write('\n')




def validate(val_loader, model, criterion, args, epoch=None, is_final_epoch=False):
    batch_time = AverageMeter('Time', ':6.3f')
    losses = AverageMeter('Loss', ':.4e')
    top1 = AverageMeter('Acc@1', ':6.2f')
    top5 = AverageMeter('Acc@5', ':6.2f')
    progress = ProgressMeter(
        len(val_loader),
        [batch_time, losses, top1, top5],
        prefix='Test: ')

    # switch to evaluate mode
    model.eval()
    all_preds = []
    all_targets = []

    with torch.no_grad():
        end = time.time()
        for i, (images, target) in enumerate(val_loader):
            if args.gpu is not None:
                images = images.cuda(args.gpu, non_blocking=True)
                target = target.cuda(args.gpu, non_blocking=True)
            if torch.cuda.is_available():
                target = target.cuda(args.gpu, non_blocking=True)

            # compute output
            output = model(images)
            loss = criterion(output, target)
            _, preds = torch.max(output, 1)
            all_preds.extend(preds.cpu().numpy())
            all_targets.extend(target.cpu().numpy())
    
            # measure accuracy and record loss
            acc1, acc5 = accuracy(output, target, topk=(1, 5))
            losses.update(loss.item(), images.size(0))
            top1.update(acc1[0], images.size(0))
            top5.update(acc5[0], images.size(0))

            # measure elapsed time
            batch_time.update(time.time() - end)
            end = time.time()

            if i % args.print_freq == 0:
                progress.display(i)

        # TODO: this should also be done with the ProgressMeter
        print(' * Acc@1 {top1.avg:.3f} Acc@5 {top5.avg:.3f}'
              .format(top1=top1, top5=top5))
    
    # Now you have all predictions and targets for the entire validation set
    # You can now calculate and save the classification report
    y_true = all_targets
    y_pred = all_preds
    conf_matrix = confusion_matrix(y_true, y_pred)
    save_path = get_save_folder_path(args)
    print(f'savepath: {save_path}')
    compute_and_save_classification_report(all_targets, all_preds, save_path, epoch, conf_matrix, args)
    
            # After the loop over the validation data
    # Assuming you have the predictions and targets stored in y_pred and y_true
    
        # Compute the metrics
    #    report = classification_report(y_true, y_pred, output_dict=True)
    #    conf_matrix = confusion_matrix(y_true, y_pred)
    balanced_acc = balanced_accuracy_score(y_true, y_pred)
    print(f'*** balanced acc: {balanced_acc}')
    class_accs = accuracy_score(all_targets, all_preds)

        # Save the classification report and confusion matrix
    #    class_metrics_df = pd.DataFrame(report).transpose()
    #    conf_matrix_df = pd.DataFrame(conf_matrix)

    #    class_metrics_df.to_csv('per_class_metrics.csv')
    #    conf_matrix_df.to_csv('confusion_matrix.csv')

    return top1.avg, balanced_acc, class_accs


def save_checkpoint(state, is_best, folder, filename='checkpoint.pth.tar'):
    file_path = os.path.join(folder, filename)
    torch.save(state, file_path)
    if is_best:
        shutil.copyfile(file_path, os.path.join(folder, 'model_best.pth.tar'))


def sanity_check(state_dict, pretrained_weights, linear_keyword):
    """
    Linear classifier should not change any weights other than the linear layer.
    This sanity check asserts nothing wrong happens (e.g., BN stats updated).
    """
    print("=> loading '{}' for sanity check".format(pretrained_weights))
    checkpoint = torch.load(pretrained_weights, map_location="cpu")
    state_dict_pre = checkpoint['state_dict']

    for k in list(state_dict.keys()):
        # only ignore linear layer
        if '%s.weight' % linear_keyword in k or '%s.bias' % linear_keyword in k:
            continue

        # name in pretrained model
        k_pre = 'module.base_encoder.' + k[len('module.'):] \
            if k.startswith('module.') else 'module.base_encoder.' + k

        assert ((state_dict[k].cpu() == state_dict_pre[k_pre]).all()), \
            '{} is changed in linear classifier training.'.format(k)

    print("=> sanity check passed.")


class AverageMeter(object):
    """Computes and stores the average and current value"""
    def __init__(self, name, fmt=':f'):
        self.name = name
        self.fmt = fmt
        self.reset()

    def reset(self):
        self.val = 0
        self.avg = 0
        self.sum = 0
        self.count = 0

    def update(self, val, n=1):
        self.val = val
        self.sum += val * n
        self.count += n
        self.avg = self.sum / self.count

    def __str__(self):
        fmtstr = '{name} {val' + self.fmt + '} ({avg' + self.fmt + '})'
        return fmtstr.format(**self.__dict__)


class ProgressMeter(object):
    def __init__(self, num_batches, meters, prefix=""):
        self.batch_fmtstr = self._get_batch_fmtstr(num_batches)
        self.meters = meters
        self.prefix = prefix

    def display(self, batch):
        entries = [self.prefix + self.batch_fmtstr.format(batch)]
        entries += [str(meter) for meter in self.meters]
        print('\t'.join(entries))

    def _get_batch_fmtstr(self, num_batches):
        num_digits = len(str(num_batches // 1))
        fmt = '{:' + str(num_digits) + 'd}'
        return '[' + fmt + '/' + fmt.format(num_batches) + ']'


def adjust_learning_rate(optimizer, init_lr, epoch, args):
    """Decay the learning rate based on schedule"""
    cur_lr = init_lr * 0.5 * (1. + math.cos(math.pi * epoch / args.epochs))
    for param_group in optimizer.param_groups:
        param_group['lr'] = cur_lr


def accuracy(output, target, topk=(1,)):
    """Computes the accuracy over the k top predictions for the specified values of k"""
    with torch.no_grad():
        maxk = max(topk)
        batch_size = target.size(0)

        _, pred = output.topk(maxk, 1, True, True)
        pred = pred.t()
        correct = pred.eq(target.view(1, -1).expand_as(pred))

        res = []
        for k in topk:
            correct_k = correct[:k].reshape(-1).float().sum(0, keepdim=True)
            res.append(correct_k.mul_(100.0 / batch_size))
        return res


if __name__ == '__main__':
    main()
